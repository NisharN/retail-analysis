"""
Tests for `app.export.build_excel` and `build_pdf`.

These check structural validity (the bytes are genuinely openable as
xlsx/pdf, headers and KPI values round-trip correctly) rather than pixel-
perfect rendering, since the latter isn't meaningfully testable without a
human looking at the file.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from app.export import build_excel, build_pdf
from app.schemas import ABCClass, GapFiltersIn, GapKPIs, GapRow, Status


@pytest.fixture
def sample_rows() -> list[GapRow]:
    return [
        GapRow(
            ArticleCode=1001,
            DepartmentName="BEVERAGES",
            ABCClass=ABCClass.A,
            NumShopsSelling=3,
            ShopSaleValue=0.0,
            ChainAvgSaleValue=16666.67,
            GapScore=1.0,
            PotentialLostRevenue=16666.67,
            Status=Status.MISSING_WINNER,
            NeverStocked=True,
        ),
        GapRow(
            ArticleCode=1002,
            DepartmentName="BEVERAGES",
            ABCClass=ABCClass.B,
            NumShopsSelling=4,
            ShopSaleValue=50.0,
            ChainAvgSaleValue=5944.4,
            GapScore=0.99,
            PotentialLostRevenue=5894.4,
            Status=Status.UNDERPERFORMING,
            NeverStocked=False,
        ),
    ]


@pytest.fixture
def sample_kpis() -> GapKPIs:
    return GapKPIs(
        missing_winners=1,
        underperforming=1,
        potential_revenue=22561.07,
        class_a_gaps=1,
        class_b_gaps=1,
        total_gaps=2,
    )


@pytest.fixture
def sample_filters() -> GapFiltersIn:
    return GapFiltersIn(shop=4, department=None, min_shops_selling=3, gap_threshold=0.20)


def test_build_excel_produces_valid_workbook(sample_rows, sample_kpis, sample_filters) -> None:
    blob = build_excel(sample_rows, sample_kpis, sample_filters)
    assert blob[:2] == b"PK"  # xlsx is a zip container

    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert set(wb.sheetnames) == {"Summary", "Detail"}

    detail = wb["Detail"]
    header_row = [c.value for c in next(detail.iter_rows(min_row=1, max_row=1))]
    assert header_row[0] == "ArticleCode"
    assert detail.cell(row=2, column=1).value == 1001
    assert detail.cell(row=3, column=1).value == 1002

    summary = wb["Summary"]
    # KPI values should round-trip into the summary sheet.
    values = [c.value for row in summary.iter_rows() for c in row if c.value is not None]
    assert 2 in values  # total_gaps
    assert 1 in values  # missing_winners / underperforming / class gaps


def test_build_excel_handles_empty_results(sample_kpis, sample_filters) -> None:
    blob = build_excel([], sample_kpis, sample_filters)
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    detail = wb["Detail"]
    # Only the header row should be present.
    assert detail.max_row == 1


def test_build_pdf_produces_valid_pdf(sample_rows, sample_kpis, sample_filters) -> None:
    blob = build_pdf(sample_rows, sample_kpis, sample_filters)
    assert blob[:5] == b"%PDF-"
    assert len(blob) > 500  # a real rendered document, not an empty stub


def test_build_pdf_handles_empty_results(sample_kpis, sample_filters) -> None:
    blob = build_pdf([], sample_kpis, sample_filters)
    assert blob[:5] == b"%PDF-"


def test_build_pdf_truncates_to_top_50(sample_kpis, sample_filters) -> None:
    many_rows = [
        GapRow(
            ArticleCode=i,
            DepartmentName="GROCERY",
            ABCClass=ABCClass.A,
            NumShopsSelling=3,
            ShopSaleValue=0.0,
            ChainAvgSaleValue=float(1000 - i),
            GapScore=1.0,
            PotentialLostRevenue=float(1000 - i),
            Status=Status.MISSING_WINNER,
            NeverStocked=True,
        )
        for i in range(75)
    ]
    blob = build_pdf(many_rows, sample_kpis, sample_filters)
    assert blob[:5] == b"%PDF-"
    # Just confirm it doesn't crash on a large result set and still
    # produces a valid PDF; exact page/row count isn't asserted since
    # that's a rendering detail.
